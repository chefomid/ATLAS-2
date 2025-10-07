# common.py
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

# ---------- Shared small helpers ----------

def next_output_path(base_dir: Path) -> Path:
    n = 1
    while True:
        p = base_dir / f"RAS_ALG_Output({n}).xlsx"
        if not p.exists():
            return p
        n += 1

def autosize_columns(ws, col_indices, last_row=None, currency_cols=None,
                     padding=2, min_width=9, max_width=60):
    if last_row is None:
        last_row = ws.max_row
    currency_cols = set(currency_cols or [])
    for c in col_indices:
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=c, max_col=c):
            v = row[0].value
            if v is None:
                s = ""
            elif isinstance(v, (int, float)):
                s = f"${v:,.2f}" if c in currency_cols else (str(int(v)) if float(v).is_integer() else str(v))
            else:
                s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(c)].width = max(min_width, min(max_len + padding, max_width))

def unique_ordered(series: pd.Series):
    seen, out = set(), []
    for v in series.astype(str):
        v = v.strip()
        if v and v not in seen:
            seen.add(v); out.append(v)
    return out

def read_sheet_any(path: Path, candidates: list[str]) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    for name in candidates:
        if name in xls.sheet_names:
            return pd.read_excel(xls, sheet_name=name)
    raise ValueError(f"None of the sheets found: {', '.join(candidates)}")

# ---------- Generic writer (used by both modules) ----------

def write_matrix(out_path: Path,
                 title: str,
                 locs: list[str],
                 covs: list[str],
                 matrix2d,                 # list[list[float]] or numpy array in locs×covs order
                 row_totals: list[float],  # per-loc row totals
                 col_totals: list[float],  # per-coverage column totals
                 meta_by_loc: dict,        # {loc: {key: value}}
                 meta_schema: list[tuple[str, str]]  # [(key, header), ...]
                 ):
    """
    Renders a matrix at A1 with optional meta columns, bottom totals row,
    currency formatting, autosized columns, and a README sheet.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "MATRIX"

    header = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_h = PatternFill("solid", fgColor="DDDDDD")
    fill_t = PatternFill("solid", fgColor="F2F2F2")
    try:
        cur = NamedStyle(name="Currency2", number_format='"$"#,##0.00')
        wb.add_named_style(cur)
    except Exception:
        pass

    # Include only meta fields that actually have values anywhere
    include_keys = []
    for key, _hdr in meta_schema:
        if any(meta_by_loc.get(l, {}).get(key, "") for l in locs):
            include_keys.append(key)
    key_to_header = {k: h for k, h in meta_schema}

    # Layout
    rH = 1
    rFirst = rH + 1
    c = 1
    cLoc = c; c += 1
    meta_cols = {}
    for k in include_keys:
        meta_cols[k] = c
        c += 1
    cFirstCov = c
    cLastCov  = cFirstCov + len(covs) - 1 if covs else cFirstCov - 1
    cRowTot   = cLastCov + 1 if covs else cFirstCov

    # Header row
    ws.cell(rH, cLoc, "Loc #")
    for k in include_keys:
        ws.cell(rH, meta_cols[k], key_to_header[k])
    for j, cov in enumerate(covs):
        ws.cell(rH, cFirstCov + j, cov)
    ws.cell(rH, cRowTot, "Total")
    for cc in range(cLoc, cRowTot + 1):
        cell = ws.cell(rH, cc)
        cell.font = header; cell.alignment = center; cell.fill = fill_h; cell.border = border

    # Data rows
    for i, loc in enumerate(locs):
        r = rFirst + i
        ws.cell(r, cLoc, loc).border = border
        meta = meta_by_loc.get(loc, {})
        for k in include_keys:
            ws.cell(r, meta_cols[k], meta.get(k, "")).border = border
        for j, _cov in enumerate(covs):
            val = float(matrix2d[i][j]) if len(covs) else 0.0
            cell = ws.cell(r, cFirstCov + j, val)
            cell.style = "Currency2"; cell.border = border
        rt = float(row_totals[i]) if i < len(row_totals) else 0.0
        rc = ws.cell(r, cRowTot, rt); rc.style = "Currency2"; rc.border = border

    # Bottom totals row
    rTotals = rFirst + len(locs)
    ws.cell(rTotals, cLoc, "Total").font = header
    ws.cell(rTotals, cLoc).fill = fill_t; ws.cell(rTotals, cLoc).border = border
    for k in include_keys:
        ws.cell(rTotals, meta_cols[k], "").border = border

    grand = 0.0
    for j in range(len(covs)):
        v = float(col_totals[j]) if j < len(col_totals) else 0.0
        grand += v
        cell = ws.cell(rTotals, cFirstCov + j, v)
        cell.style = "Currency2"; cell.border = border; cell.font = header; cell.fill = fill_t
    gcell = ws.cell(rTotals, cRowTot, grand)
    gcell.style = "Currency2"; gcell.border = border; gcell.font = header; gcell.fill = fill_t

    # Autosize
    currency_cols = set(range(cFirstCov, cRowTot + 1))
    autosize_columns(ws, range(cLoc, cRowTot + 1), last_row=rTotals, currency_cols=currency_cols)

    # README
    rm = wb.create_sheet("READ_ME")
    rm["A1"] = title
    rm["A1"].font = header
    rm["A3"] = "Generated by RAS/TIV distributor."

    wb.save(out_path)
