from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

# ---------- small shared helpers ----------
def next_output_path(base_dir: Path) -> Path:
    n = 1
    while True:
        p = base_dir / f"TIV_Weighted_Matrix({n}).xlsx"
        if not p.exists():
            return p
        n += 1

def autosize_columns(ws, col_indices, last_row=None, currency_cols=None,
                     padding=2, min_width=9, max_width=60):
    if last_row is None:
        last_row = ws.max_row
    currency_cols = set(currency_cols or [])
    for c in col_indices:
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=c, max_col=c):
            v = row[0].value
            if v is None:
                s = ""
            elif isinstance(v, (int, float)):
                s = f"${v:,.2f}" if c in currency_cols else (str(int(v)) if float(v).is_integer() else str(v))
            else:
                s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(c)].width = max(min_width, min(max_len + padding, max_width))

def unique_ordered(series: pd.Series):
    seen, out = set(), []
    for v in series.astype(str):
        v = v.strip()
        if v and v.lower() not in ("nan", "none") and v not in seen:
            seen.add(v); out.append(v)
    return out

def read_sheet_any(path: Path, candidates: list[str]) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    for name in candidates:
        if name in xls.sheet_names:
            return pd.read_excel(xls, sheet_name=name)
    raise ValueError(f"None of the sheets found: {', '.join(candidates)}")

# ---------- TIV specifics ----------
SHEET_TIV = "TIV Weighted Dist. - INPUT"

COV_NAMES    = ["Coverage Type", "Coverage/Expense", "Coverage"]
PREM_NAMES   = ["Premium Amount", "Premium", "Premium Total"]
LOC_NAMES    = ["Loc #", "Location #", "Loc"]
ENT_NAMES    = ["Enitity Name", "Entity Name"]
STREET_NAMES = ["Street", "Address 1", "Address1"]
CITY_NAMES   = ["City"]
STATE_NAMES  = ["State", "ST"]
ZIP_NAMES    = ["Zip-Code", "Zip-code", "Zip Code", "Zip", "Postal Code"]
TIV_NAMES    = [
    "Insurable Value", "Total Insured Value",
    "TIV", "TIV ($)", "TIV Amount", "TIV Value", "TIV USD", "TIV Total",
    "Insured Value", "Replacement Cost",
]

def pick_col(df, candidates, default=None):
    df.columns = [str(c).replace("\xa0", " ").strip() for c in df.columns]
    def norm(s: str) -> str: return "".join(ch for ch in str(s).lower() if ch.isalnum())
    norm_map = {norm(c): c for c in df.columns}
    for cand in candidates:
        key = norm(cand)
        if key in norm_map:
            return norm_map[key]
    for cand in candidates:
        key = norm(cand)
        for k, orig in norm_map.items():
            if key in k:
                return orig
    return default

def _clean_text(x):
    if pd.isna(x):
        return ""
    s = str(x).replace("\xa0", " ").strip()
    return "" if s.lower() in ("nan", "none") else s

def _normalize_loc(x):
    s = _clean_text(x)
    if not s:
        return np.nan
    # turn "1.0" -> "1"
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except Exception:
        pass
    return s

def load_tiv_sheet(path: Path) -> tuple[pd.DataFrame, dict]:
    df = read_sheet_any(path, [SHEET_TIV])
    df.columns = [str(c).strip() for c in df.columns]

    col_cov  = pick_col(df, COV_NAMES)
    col_pre  = pick_col(df, PREM_NAMES)
    col_loc  = pick_col(df, LOC_NAMES)
    col_ent  = pick_col(df, ENT_NAMES)
    col_st   = pick_col(df, STREET_NAMES)
    col_city = pick_col(df, CITY_NAMES)
    col_state= pick_col(df, STATE_NAMES)
    col_zip  = pick_col(df, ZIP_NAMES)
    col_tiv  = pick_col(df, TIV_NAMES)

    missing = [n for n, c in [
        ("Coverage", col_cov),
        ("Premium Amount", col_pre),
        ("Loc #", col_loc),
        ("TIV/Insurable Value", col_tiv),
    ] if c is None]
    if missing:
        raise ValueError(
            f"Missing required column(s) on '{SHEET_TIV}': {', '.join(missing)}\n"
            f"Headers found: {', '.join(map(str, df.columns))}"
        )

    # Clean/normalize
    df[col_cov] = df[col_cov].map(_clean_text)
    df[col_loc] = df[col_loc].map(_normalize_loc)  # <- removes .0, blanks -> NaN
    df[col_pre] = pd.to_numeric(df[col_pre], errors="coerce").fillna(0.0)
    df[col_tiv] = pd.to_numeric(df[col_tiv], errors="coerce").fillna(0.0)

    for c in [col_ent, col_st, col_city, col_state, col_zip]:
        if c is not None:
            df[c] = df[c].map(_clean_text)

    cols = dict(cov=col_cov, pre=col_pre, loc=col_loc, ent=col_ent, street=col_st,
                city=col_city, state=col_state, zip=col_zip, tiv=col_tiv)
    return df, cols

def build_loc_meta_tiv(df: pd.DataFrame, cols: dict) -> dict:
    meta = {}
    for _, r in df.iterrows():
        loc = r[cols["loc"]]
        if pd.isna(loc):
            continue
        ent  = _clean_text(r[cols["ent"]])   if cols["ent"]   in df.columns else ""
        st   = _clean_text(r[cols["street"]])if cols["street"]in df.columns else ""
        city = _clean_text(r[cols["city"]])  if cols["city"]  in df.columns else ""
        state= _clean_text(r[cols["state"]]) if cols["state"] in df.columns else ""
        zipc = _clean_text(r[cols["zip"]])   if cols["zip"]   in df.columns else ""
        loc = str(loc)
        if loc not in meta:
            meta[loc] = {"entity": ent, "street": st, "city": city, "state": state, "zip": zipc}
        else:
            m = meta[loc]
            if not m["entity"] and ent: m["entity"] = ent
            if not m["street"] and st:  m["street"]  = st
            if not m["city"] and city:  m["city"]    = city
            if not m["state"] and state:m["state"]   = state
            if not m["zip"] and zipc:   m["zip"]     = zipc
    return meta

def allocate_cents_for_coverage(premium_total, tiv_by_loc_map, loc_list):
    """
    Distribute premium_total across loc_list using tiv_by_loc_map[loc] weights.
    Exact cents via largest remainders. Returns dict loc->amount (float dollars).
    """
    loc_list = list(loc_list)
    n = len(loc_list)
    if n == 0:
        return {}
    tivs = np.array([float(tiv_by_loc_map.get(l, 0.0)) for l in loc_list], dtype=float)
    total_tiv = tivs.sum()

    if premium_total <= 0:
        return {l: 0.0 for l in loc_list}

    if total_tiv <= 0:
        weights = np.ones(n) / n
    else:
        weights = tivs / total_tiv

    cents_total = int(round(premium_total * 100))
    raw = weights * cents_total
    floor = np.floor(raw).astype(int)
    rema = raw - floor
    remain = cents_total - int(floor.sum())

    order = np.argsort(-rema)
    add = np.zeros(n, dtype=int)
    if remain > 0:
        add[order[:remain]] = 1

    cents = floor + add
    # safety
    assert int(cents.sum()) == cents_total
    return {loc_list[i]: float(cents[i]) / 100.0 for i in range(n)}

def build_tiv_matrix(df: pd.DataFrame, cols: dict):
    # Real location labels (non-blank, normalized)
    locs = unique_ordered(df[cols["loc"]].dropna())

    # GLOBAL TIV by location (default vector)
    tiv_global_map = df.dropna(subset=[cols["loc"]]).groupby(cols["loc"])[cols["tiv"]].sum().to_dict()
    # Ensure every loc has a key
    for l in locs:
        tiv_global_map.setdefault(l, 0.0)

    # Premium totals per coverage
    covs = unique_ordered(df[cols["cov"]])
    premium_by_cov = df.groupby(cols["cov"])[cols["pre"]].sum().to_dict()

    # Optional coverage-specific TIVs (if sheet supplies multiple locs per coverage)
    tiv_by_cov_loc = df.dropna(subset=[cols["loc"]]).groupby([cols["cov"], cols["loc"]])[cols["tiv"]].sum().to_dict()

    # Build matrix loc×cov
    mat = {str(l): {c: 0.0 for c in covs} for l in locs}

    for cov in covs:
        P = float(premium_by_cov.get(cov, 0.0))

        # Pull TIVs for this coverage if it spans multiple locs with positive TIVs
        cov_locs = [l for l in locs if (cov, l) in tiv_by_cov_loc and tiv_by_cov_loc[(cov, l)] > 0]
        if len(cov_locs) >= 2:
            tiv_map = {l: tiv_by_cov_loc.get((cov, l), 0.0) for l in cov_locs}
            alloc_locs = cov_locs
        else:
            # Fallback: use the GLOBAL TIV vector (spreads over all locs)
            tiv_map = tiv_global_map
            alloc_locs = locs

        alloc = allocate_cents_for_coverage(P, tiv_map, alloc_locs)
        for l in alloc_locs:
            mat[str(l)][cov] = alloc.get(l, 0.0)

    # Totals
    row_vec = [sum(mat[str(l)][c] for c in covs) for l in locs]
    col_vec = [float(premium_by_cov.get(c, 0.0)) for c in covs]

    # Meta
    loc_meta = build_loc_meta_tiv(df, cols)
    return [str(l) for l in locs], covs, mat, row_vec, col_vec, loc_meta

def write_matrix_generic(out_path: Path, locs, covs, matrix_dict, row_totals_vec, col_totals_vec, meta_by_loc: dict):
    # Convert dict -> 2D array in locs×covs order
    matrix_2d = np.array([[matrix_dict[l][c] for c in covs] for l in locs], dtype=float)

    wb = Workbook()
    ws = wb.active
    ws.title = "MATRIX"

    header = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_h = PatternFill("solid", fgColor="DDDDDD")
    fill_t = PatternFill("solid", fgColor="F2F2F2")
    try:
        wb.add_named_style(NamedStyle(name="Currency2", number_format='"$"#,##0.00'))
    except Exception:
        pass

    include_entity = any(meta_by_loc.get(l, {}).get("entity", "") for l in locs)
    include_street = any(meta_by_loc.get(l, {}).get("street", "") for l in locs)
    include_city   = any(meta_by_loc.get(l, {}).get("city",   "") for l in locs)
    include_state  = any(meta_by_loc.get(l, {}).get("state",  "") for l in locs)
    include_zip    = any(meta_by_loc.get(l, {}).get("zip",    "") for l in locs)

    rH = 1; rFirst = rH + 1; c = 1
    cLoc = c; c += 1
    cEnt = c if include_entity else None;  c += int(include_entity)
    cSt  = c if include_street else None;  c += int(include_street)
    cCity= c if include_city   else None;  c += int(include_city)
    cStt = c if include_state  else None;  c += int(include_state)
    cZip = c if include_zip    else None;  c += int(include_zip)
    cFirstCov = c
    cLastCov  = cFirstCov + len(covs) - 1 if covs else cFirstCov - 1
    cRowTot   = cLastCov + 1 if covs else cFirstCov

    ws.cell(rH, cLoc, "Loc #")
    if include_entity: ws.cell(rH, cEnt, "Entity Name")
    if include_street: ws.cell(rH, cSt,  "Street")
    if include_city:   ws.cell(rH, cCity,"City")
    if include_state:  ws.cell(rH, cStt, "State")
    if include_zip:    ws.cell(rH, cZip, "Zip")
    for j, cov in enumerate(covs):
        ws.cell(rH, cFirstCov + j, cov)
    ws.cell(rH, cRowTot, "Total")
    for cc in range(cLoc, cRowTot + 1):
        cell = ws.cell(rH, cc)
        cell.font = header; cell.alignment = center; cell.fill = fill_h; cell.border = border

    for i, loc in enumerate(locs):
        r = rFirst + i
        ws.cell(r, cLoc, loc).border = border
        m = meta_by_loc.get(loc, {})
        if include_entity: ws.cell(r, cEnt, m.get("entity","")).border = border
        if include_street: ws.cell(r, cSt,  m.get("street","")).border = border
        if include_city:   ws.cell(r, cCity,m.get("city","")).border = border
        if include_state:  ws.cell(r, cStt, m.get("state","")).border = border
        if include_zip:    ws.cell(r, cZip, m.get("zip","")).border = border
        for j, cov in enumerate(covs):
            cell = ws.cell(r, cFirstCov + j, float(matrix_2d[i][j]) if len(covs) else 0.0)
            cell.style = "Currency2"; cell.border = border
        rc = ws.cell(r, cRowTot, float(row_totals_vec[i]) if i < len(row_totals_vec) else 0.0)
        rc.style = "Currency2"; rc.border = border

    rTotals = rFirst + len(locs)
    ws.cell(rTotals, cLoc, "Total").font = header
    ws.cell(rTotals, cLoc).fill = fill_t; ws.cell(rTotals, cLoc).border = border
    if include_entity: ws.cell(rTotals, cEnt, "").border = border
    if include_street: ws.cell(rTotals, cSt,  "").border = border
    if include_city:   ws.cell(rTotals, cCity,"").border = border
    if include_state:  ws.cell(rTotals, cStt, "").border = border
    if include_zip:    ws.cell(rTotals, cZip, "").border = border

    grand = 0.0
    for j in range(len(covs)):
        v = float(col_totals_vec[j]) if j < len(col_totals_vec) else 0.0
        grand += v
        cell = ws.cell(rTotals, cFirstCov + j, v)
        cell.style = "Currency2"; cell.border = border; cell.font = header; cell.fill = fill_t
    gcell = ws.cell(rTotals, cRowTot, grand)
    gcell.style = "Currency2"; gcell.border = border; gcell.font = header; gcell.fill = fill_t

    currency_cols = set(range(cFirstCov, cRowTot + 1))
    autosize_columns(ws, range(cLoc, cRowTot + 1), last_row=rTotals, currency_cols=currency_cols)

    rm = wb.create_sheet("READ_ME")
    rm["A1"] = "TIV Weighted Distribution"; rm["A1"].font = header
    rm["A3"] = "Premium per coverage distributed by GLOBAL TIV share unless coverage-specific TIVs for multiple locs are provided."
    rm["A4"] = "Rounding: exact cents per coverage."

    wb.save(out_path)

# ---------- public entry ----------
def build_tiv(path_str: str) -> Path:
    inp = Path(path_str)
    df, cols = load_tiv_sheet(inp)
    locs, covs, mat_dict, row_vec, col_vec, loc_meta = build_tiv_matrix(df, cols)
    out_path = next_output_path(inp.parent)
    write_matrix_generic(out_path, locs, covs, mat_dict, row_vec, col_vec, loc_meta)
    return out_path
