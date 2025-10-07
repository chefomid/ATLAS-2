# ras_alg_simple.py
# Minimal UI to build:
#  - Skeleton matrix (zeros interior), or
#  - Balanced matrix via IPF (RAS) + exact two-decimal rounding (no penny drift)
# Reads an Excel template (sheet "INPUT") and writes a new workbook "RAS_ALG_Output(n).xlsx"
# Requirements: pandas, numpy, openpyxl (Tkinter is included with standard Python on Windows)

import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

# ------------------ Column names (must match your template headers) ------------------
COL_LOC  = "Loc #"              # Column A
COL_ROW  = "Premium Total"      # Column D
COL_COV  = "Coverage/Expense"   # Column G
COL_COL  = "Total"              # Column H

# Optional columns (exact header from your template)
COL_ENT  = "Enitity Name"       # Column B (intentional spelling per template)
COL_ADDR = "Address"            # Column C

# ------------------ Autosize helper ------------------
def autosize_columns(ws, col_indices, last_row=None, currency_cols=None,
                     padding=2, min_width=9, max_width=60):
    """
    Auto-size Excel columns based on longest rendered text in those columns.
    currency_cols: set of 1-based column indices that should be measured as $#,##0.00.
    """
    if last_row is None:
        last_row = ws.max_row
    currency_cols = set(currency_cols or [])
    for c in col_indices:
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=c, max_col=c):
            v = row[0].value
            if v is None:
                s = ""
            elif isinstance(v, (int, float)):
                s = f"${v:,.2f}" if c in currency_cols else (str(int(v)) if float(v).is_integer() else str(v))
            else:
                s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(c)].width = max(min_width, min(max_len + padding, max_width))

# ------------------ IO + preprocessing ------------------
def load_template(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="INPUT")
    df.columns = [str(c).strip() for c in df.columns]
    # Ensure columns exist
    for c in (COL_LOC, COL_COV, COL_ROW, COL_COL, COL_ENT, COL_ADDR):
        if c not in df.columns:
            df[c] = np.nan
    # Clean
    df[COL_LOC]  = df[COL_LOC].astype(str).str.strip()
    df[COL_COV]  = df[COL_COV].astype(str).str.strip()
    df[COL_ENT]  = df[COL_ENT].astype(str).str.strip()
    df[COL_ADDR] = df[COL_ADDR].astype(str).str.strip()
    df[COL_ROW]  = pd.to_numeric(df[COL_ROW], errors="coerce").fillna(0.0)
    df[COL_COL]  = pd.to_numeric(df[COL_COL], errors="coerce").fillna(0.0)
    # Treat literal "nan" strings as blanks
    for col in (COL_LOC, COL_COV, COL_ENT, COL_ADDR):
        df.loc[df[col].str.lower() == "nan", col] = ""
    return df

def unique_ordered(series: pd.Series) -> list[str]:
    seen, out = set(), []
    for v in series.astype(str):
        v = v.strip()
        if v and v not in seen:
            seen.add(v); out.append(v)
    return out

def build_loc_metadata(df: pd.DataFrame) -> dict:
    """
    For each Loc #, fetch first non-blank Entity Name and Address encountered.
    If duplicates conflict, first non-blank wins (stable order).
    """
    meta = {}
    for _, row in df.iterrows():
        loc = str(row[COL_LOC]).strip()
        if not loc:
            continue
        ent  = (str(row[COL_ENT]).strip()  if pd.notna(row[COL_ENT])  else "")
        addr = (str(row[COL_ADDR]).strip() if pd.notna(row[COL_ADDR]) else "")
        if loc not in meta:
            meta[loc] = {"entity": ent, "address": addr}
        else:
            if not meta[loc]["entity"] and ent:
                meta[loc]["entity"] = ent
            if not meta[loc]["address"] and addr:
                meta[loc]["address"] = addr
    return meta

def aggregates(df: pd.DataFrame):
    locs = unique_ordered(df[COL_LOC])
    covs = unique_ordered(df[COL_COV])
    row_totals = df.groupby(COL_LOC, dropna=False)[COL_ROW].sum().to_dict()
    col_totals = df.groupby(COL_COV, dropna=False)[COL_COL].sum().to_dict()
    loc_meta = build_loc_metadata(df)
    return locs, covs, row_totals, col_totals, loc_meta

# ------------------ IPF / RAS ------------------
def ipf(row_targets: np.ndarray, col_targets: np.ndarray, seed: np.ndarray | None = None,
        max_iter=5000, tol=1e-10):
    """Iterative proportional fitting to match row and column targets in real numbers."""
    rT = np.asarray(row_targets, dtype=float)
    cT = np.asarray(col_targets, dtype=float)
    nR, nC = len(rT), len(cT)
    if seed is None:
        seed = np.outer((rT > 0).astype(float), (cT > 0).astype(float))
        if seed.sum() == 0:
            seed = np.ones((nR, nC), dtype=float)
    X = seed.copy()
    # Rough scale once
    rs = X.sum(axis=1, keepdims=True); rs[rs == 0] = 1.0
    X *= (rT / rs.squeeze())[:, None]
    cs = X.sum(axis=0, keepdims=True); cs[cs == 0] = 1.0
    X *= (cT / cs.squeeze())[None, :]
    # Iterate
    for _ in range(max_iter):
        rs = X.sum(axis=1, keepdims=True); rs[rs == 0] = 1.0
        X *= (rT / rs.squeeze())[:, None]
        cs = X.sum(axis=0, keepdims=True); cs[cs == 0] = 1.0
        X *= (cT / cs.squeeze())[None, :]
        if np.allclose(X.sum(axis=1), rT, atol=tol) and np.allclose(X.sum(axis=0), cT, atol=tol):
            break
    return X

# ------------------ Exact two-decimal rounding ------------------
def round_matrix_exact_cents(X: np.ndarray, row_targets: list[float], col_targets: list[float]) -> np.ndarray:
    """
    Convert a real-valued matrix to dollars with two decimals such that
    row sums and column sums match the given targets exactly.
    Steps: convert to cents → floor → allocate remaining cents using largest remainders,
    while respecting both row and column deficits.
    """
    rT_c = np.rint(np.asarray(row_targets, float) * 100).astype(int)
    cT_c = np.rint(np.asarray(col_targets, float) * 100).astype(int)

    Xc_real = np.asarray(X, float) * 100.0
    Xc = np.floor(Xc_real).astype(int)
    R = Xc_real - Xc

    row_def = rT_c - Xc.sum(axis=1)
    col_def = cT_c - Xc.sum(axis=0)

    # Global remainder ordering (descending)
    idx = [(i, j) for i in range(Xc.shape[0]) for j in range(Xc.shape[1])]
    idx.sort(key=lambda ij: R[ij[0], ij[1]], reverse=True)

    # Greedy bipartite allocation
    for i, j in idx:
        if row_def[i] > 0 and col_def[j] > 0:
            take = min(row_def[i], col_def[j])
            Xc[i, j] += take
            row_def[i] -= take
            col_def[j] -= take
        if row_def.sum() == 0 and col_def.sum() == 0:
            break

    return Xc.astype(float) / 100.0

# ------------------ Excel writer ------------------
def write_matrix(out_path: Path, locs, covs, matrix, row_totals, col_totals, loc_meta):
    wb = Workbook()
    ws = wb.active
    ws.title = "MATRIX"

    header = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_h = PatternFill("solid", fgColor="DDDDDD")
    fill_t = PatternFill("solid", fgColor="F2F2F2")
    try:
        cur = NamedStyle(name="Currency2", number_format='"$"#,##0.00')
        wb.add_named_style(cur)
    except Exception:
        pass

    # Anchor at A1
    rH = 1
    rFirstData = rH + 1
    # Columns:
    # A: Loc #, B: Entity Name (optional), C: Address (optional), D..: Coverages, last: Total
    cLoc = 1
    include_entity  = any(loc_meta.get(l, {}).get("entity", "") for l in locs)
    include_address = any(loc_meta.get(l, {}).get("address", "") for l in locs)
    cEnt = cLoc + 1 if include_entity else None
    cAdr = (cEnt + 1) if include_address and include_entity else ((cLoc + 1) if include_address else None)
    cFirstCov = (cLoc + 1) + int(include_entity) + int(include_address)

    nL, nC = len(locs), len(covs)
    cLastCov = cFirstCov + (nC - 1) if nC > 0 else cFirstCov - 1
    cRowTot  = cLastCov + 1 if nC > 0 else cFirstCov

    # Header row
    ws.cell(rH, cLoc, "Loc #")
    if include_entity:
        ws.cell(rH, cEnt, "Entity Name")
    if include_address:
        ws.cell(rH, cAdr, "Address")
    for j, cov in enumerate(covs):
        ws.cell(rH, cFirstCov + j, cov)
    ws.cell(rH, cRowTot, "Total")

    for c in range(cLoc, cRowTot + 1):
        cell = ws.cell(rH, c)
        cell.font = header
        cell.alignment = center
        cell.fill = fill_h
        cell.border = border

    # Data rows
    for i, loc in enumerate(locs):
        r = rFirstData + i
        ws.cell(r, cLoc, loc).border = border
        meta = loc_meta.get(loc, {"entity": "", "address": ""})
        if include_entity:
            ws.cell(r, cEnt, meta.get("entity", "")).border = border
        if include_address:
            ws.cell(r, cAdr, meta.get("address", "")).border = border
        # body
        for j in range(nC):
            val = float(matrix[i][j]) if nC else 0.0
            cell = ws.cell(r, cFirstCov + j, val)
            cell.style = "Currency2"
            cell.border = border
        # row total (from inputs)
        rt = float(row_totals.get(loc, 0.0))
        rc = ws.cell(r, cRowTot, rt)
        rc.style = "Currency2"
        rc.border = border

    # Bottom totals row (after last location)
    rTotals = rFirstData + max(nL, 0)
    ws.cell(rTotals, cLoc, "Total").font = header
    ws.cell(rTotals, cLoc).fill = fill_t
    ws.cell(rTotals, cLoc).border = border
    if include_entity:
        ws.cell(rTotals, cEnt, "").border = border
    if include_address:
        ws.cell(rTotals, cAdr, "").border = border

    grand = 0.0
    for j, cov in enumerate(covs):
        v = float(col_totals.get(cov, 0.0))
        cell = ws.cell(rTotals, cFirstCov + j, v)
        cell.style = "Currency2"
        cell.border = border
        cell.font = header
        cell.fill = fill_t
        grand += v
    gcell = ws.cell(rTotals, cRowTot, grand)
    gcell.style = "Currency2"
    gcell.border = border
    gcell.font = header
    gcell.fill = fill_t

    # ---------- Auto-size columns (AFTER all content is written) ----------
    currency_cols = set(range(cFirstCov, cRowTot + 1))  # all coverage cols + far-right Total
    autosize_columns(
        ws,
        col_indices=range(cLoc, cRowTot + 1),
        last_row=rTotals,                  # scan through bottom totals row
        currency_cols=currency_cols,
        padding=2,
        min_width=9,
        max_width=60
    )

    # README
    rm = wb.create_sheet("READ_ME")
    rm["A1"] = "RAS Matrix"
    rm["A1"].font = header
    rm["A3"] = "Optional columns included if present: 'Enitity Name', 'Address'."
    rm["A5"] = "Modes:"
    rm["A6"] = "- Skeleton: zeros interior"
    rm["A7"] = "- Balanced: IPF (RAS) + exact two-decimal rounding (cents apportionment)"

    wb.save(out_path)

# ------------------ Output naming ------------------
def next_output_path(base_dir: Path) -> Path:
    n = 1
    while True:
        p = base_dir / f"RAS_ALG_Output({n}).xlsx"
        if not p.exists():
            return p
        n += 1

# ------------------ Build pipeline ------------------
def build_from_file(path_str: str, mode: str):
    inp = Path(path_str)
    if not inp.exists():
        raise FileNotFoundError(inp)
    df = load_template(inp)
    locs, covs, row_totals, col_totals, loc_meta = aggregates(df)

    nL, nC = len(locs), len(covs)
    body = np.zeros((max(nL, 0), max(nC, 0)), dtype=float)

    if mode == "skeleton" or nL == 0 or nC == 0:
        M = body
    else:
        row_vec = [float(row_totals.get(loc, 0.0)) for loc in locs]
        col_vec = [float(col_totals.get(cov, 0.0)) for cov in covs]
        rt = np.array(row_vec, dtype=float)
        ct = np.array(col_vec, dtype=float)

        # Seed proportional to outer product of targets (masked by zero rows/cols)
        seed = np.outer(np.where(rt > 0, rt, 0), np.where(ct > 0, ct, 0))
        seed = seed / seed.sum() * max(rt.sum(), 1.0) if seed.sum() > 0 else None

        # IPF to match real-valued margins
        M_real = ipf(rt, ct, seed=seed)
        # Exact 2-decimal rounding with cents apportionment
        M = round_matrix_exact_cents(M_real, row_vec, col_vec)

    out_path = next_output_path(inp.parent)
    write_matrix(out_path, locs, covs, M, row_totals, col_totals, loc_meta)
    return out_path

# ------------------ Minimal UI ------------------
def main():
    root = tk.Tk()
    root.title("RAS Matrix Builder")
    root.geometry("540x220")
    root.resizable(False, False)

    tk.Label(
        root,
        text="Select template (.xlsx) with sheet 'INPUT'\n"
             "A: Loc # | B: Enitity Name (opt) | C: Address (opt) | D: Premium Total | G: Coverage/Expense | H: Total"
    ).pack(pady=8)

    mode_var = tk.StringVar(value="balanced")
    frame = tk.Frame(root); frame.pack()
    tk.Radiobutton(frame, text="Balanced (RAS + exact cents)", variable=mode_var, value="balanced").grid(row=0, column=0, padx=6)
    tk.Radiobutton(frame, text="Skeleton (zeros inside)",       variable=mode_var, value="skeleton").grid(row=0, column=1, padx=6)

    def on_pick():
        fp = filedialog.askopenfilename(title="Choose template Excel", filetypes=[("Excel Files", "*.xlsx")])
        if not fp:
            return
        try:
            outp = build_from_file(fp, mode_var.get())
            messagebox.showinfo("Done", f"Matrix saved:\n{outp}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    tk.Button(root, text="Build Matrix", width=20, command=on_pick).pack(pady=12)
    tk.Label(root, text="Output: RAS_ALG_Output(n).xlsx", fg="#555").pack()

    root.mainloop()

if __name__ == "__main__":
    main()
