from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

# ---------- helpers ----------
def next_output_path(base_dir: Path) -> Path:
    n = 1
    while True:
        p = base_dir / f"RAS_ALG_Output({n}).xlsx"
        if not p.exists():
            return p
        n += 1

def autosize_columns(ws, col_indices, last_row=None, currency_cols=None,
                     padding=2, min_width=9, max_width=60):
    if last_row is None:
        last_row = ws.max_row
    currency_cols = set(currency_cols or [])
    for c in col_indices:
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=c, max_col=c):
            v = row[0].value
            if v is None:
                s = ""
            elif isinstance(v, (int, float)):
                s = f"${v:,.2f}" if c in currency_cols else (str(int(v)) if float(v).is_integer() else str(v))
            else:
                s = str(v)
            max_len = max(max_len, len(s))
        ws.column_dimensions[get_column_letter(c)].width = max(min_width, min(max_len + padding, max_width))

def unique_ordered(series: pd.Series):
    seen, out = set(), []
    for v in series.astype(str):
        v = v.strip()
        if v and v not in seen:
            seen.add(v); out.append(v)
    return out

def read_sheet_any(path: Path, candidates: list[str]) -> pd.DataFrame:
    xls = pd.ExcelFile(path)
    for name in candidates:
        if name in xls.sheet_names:
            return pd.read_excel(xls, sheet_name=name)
    raise ValueError(f"None of the sheets found: {', '.join(candidates)}")

# ---------- RAS specifics ----------
RAS_SHEETS = ["RAS Algorithm - INPUT", "INPUT"]
COL_LOC  = "Loc #"
COL_ROW  = "Premium Total"        # D
COL_COV  = "Coverage/Expense"     # G
COL_COL  = "Total"                # H
COL_ENT  = "Enitity Name"         # B
COL_ADDR = "Address"              # C

def load_ras_sheet(path: Path) -> pd.DataFrame:
    df = read_sheet_any(path, RAS_SHEETS)
    df.columns = [str(c).strip() for c in df.columns]
    for c in (COL_LOC, COL_COV, COL_ROW, COL_COL, COL_ENT, COL_ADDR):
        if c not in df.columns:
            df[c] = np.nan
    df[COL_LOC]  = df[COL_LOC].astype(str).str.strip()
    df[COL_COV]  = df[COL_COV].astype(str).str.strip()
    df[COL_ENT]  = df[COL_ENT].astype(str).str.strip()
    df[COL_ADDR] = df[COL_ADDR].astype(str).str.strip()
    df[COL_ROW]  = pd.to_numeric(df[COL_ROW], errors="coerce").fillna(0.0)
    df[COL_COL]  = pd.to_numeric(df[COL_COL], errors="coerce").fillna(0.0)
    for col in (COL_LOC, COL_COV, COL_ENT, COL_ADDR):
        df.loc[df[col].str.lower() == "nan", col] = ""
    return df

def build_loc_meta_ras(df: pd.DataFrame) -> dict:
    meta = {}
    for _, row in df.iterrows():
        loc = str(row[COL_LOC]).strip()
        if not loc:
            continue
        ent  = str(row[COL_ENT]).strip()  if pd.notna(row[COL_ENT])  else ""
        addr = str(row[COL_ADDR]).strip() if pd.notna(row[COL_ADDR]) else ""
        if loc not in meta:
            meta[loc] = {"entity": ent, "address": addr}
        else:
            if not meta[loc]["entity"] and ent:   meta[loc]["entity"] = ent
            if not meta[loc]["address"] and addr: meta[loc]["address"] = addr
    return meta

def ipf(row_targets: np.ndarray, col_targets: np.ndarray, seed: np.ndarray | None = None,
        max_iter=5000, tol=1e-10):
    rT = np.asarray(row_targets, dtype=float)
    cT = np.asarray(col_targets, dtype=float)
    nR, nC = len(rT), len(cT)
    if seed is None:
        seed = np.outer((rT > 0).astype(float), (cT > 0).astype(float))
        if seed.sum() == 0:
            seed = np.ones((nR, nC), dtype=float)
    X = seed.copy()
    rs = X.sum(axis=1, keepdims=True); rs[rs == 0] = 1.0
    X *= (rT / rs.squeeze())[:, None]
    cs = X.sum(axis=0, keepdims=True); cs[cs == 0] = 1.0
    X *= (cT / cs.squeeze())[None, :]
    for _ in range(max_iter):
        rs = X.sum(axis=1, keepdims=True); rs[rs == 0] = 1.0
        X *= (rT / rs.squeeze())[:, None]
        cs = X.sum(axis=0, keepdims=True); cs[cs == 0] = 1.0
        X *= (cT / cs.squeeze())[None, :]
        if np.allclose(X.sum(axis=1), rT, atol=tol) and np.allclose(X.sum(axis=0), cT, atol=tol):
            break
    return X

def round_matrix_exact_cents(X: np.ndarray, row_targets: list[float], col_targets: list[float]) -> np.ndarray:
    rT_c = np.rint(np.asarray(row_targets, float) * 100).astype(int)
    cT_c = np.rint(np.asarray(col_targets, float) * 100).astype(int)
    Xc_real = np.asarray(X, float) * 100.0
    Xc = np.floor(Xc_real).astype(int)
    R = Xc_real - Xc
    row_def = rT_c - Xc.sum(axis=1)
    col_def = cT_c - Xc.sum(axis=0)
    idx = [(i, j) for i in range(Xc.shape[0]) for j in range(Xc.shape[1])]
    idx.sort(key=lambda ij: R[ij[0], ij[1]], reverse=True)
    for i, j in idx:
        if row_def[i] > 0 and col_def[j] > 0:
            take = min(row_def[i], col_def[j])
            Xc[i, j] += take
            row_def[i] -= take
            col_def[j] -= take
        if row_def.sum() == 0 and col_def.sum() == 0:
            break
    return Xc.astype(float) / 100.0

def build_ras_matrix(df: pd.DataFrame):
    locs = unique_ordered(df[COL_LOC])
    covs = unique_ordered(df[COL_COV])
    row_totals_map = df.groupby(COL_LOC, dropna=False)[COL_ROW].sum().to_dict()
    col_totals_map = df.groupby(COL_COV, dropna=False)[COL_COL].sum().to_dict()
    row_vec = [float(row_totals_map.get(l, 0.0)) for l in locs]
    col_vec = [float(col_totals_map.get(c, 0.0)) for c in covs]
    nL, nC = len(locs), len(covs)
    if nL == 0 or nC == 0:
        M = np.zeros((max(nL, 0), max(nC, 0)), dtype=float)
    else:
        rt = np.array(row_vec, dtype=float)
        ct = np.array(col_vec, dtype=float)
        seed = np.outer(np.where(rt > 0, rt, 0), np.where(ct > 0, ct, 0))
        seed = seed / seed.sum() * max(rt.sum(), 1.0) if seed.sum() > 0 else None
        M_real = ipf(rt, ct, seed=seed)
        M = round_matrix_exact_cents(M_real, row_vec, col_vec)
    return locs, covs, M, row_vec, col_vec, build_loc_meta_ras(df)

def write_matrix_generic(out_path: Path, locs, covs, matrix_2d, row_totals_vec, col_totals_vec,
                         meta_by_loc: dict):
    wb = Workbook(); ws = wb.active; ws.title = "MATRIX"
    header = Font(bold=True); center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_h = PatternFill("solid", fgColor="DDDDDD"); fill_t = PatternFill("solid", fgColor="F2F2F2")
    try:
        wb.add_named_style(NamedStyle(name="Currency2", number_format='"$"#,##0.00'))
    except Exception:
        pass

    include_entity  = any(meta_by_loc.get(l, {}).get("entity", "") for l in locs)
    include_address = any(meta_by_loc.get(l, {}).get("address", "") for l in locs)

    rH = 1; rFirst = rH + 1; c = 1
    cLoc = c; c += 1
    cEnt = c if include_entity else None;  c += int(include_entity)
    cAdr = c if include_address else None; c += int(include_address)
    cFirstCov = c
    cLastCov  = cFirstCov + len(covs) - 1 if covs else cFirstCov - 1
    cRowTot   = cLastCov + 1 if covs else cFirstCov

    ws.cell(rH, cLoc, "Loc #")
    if include_entity:  ws.cell(rH, cEnt, "Entity Name")
    if include_address: ws.cell(rH, cAdr, "Address")
    for j, cov in enumerate(covs): ws.cell(rH, cFirstCov + j, cov)
    ws.cell(rH, cRowTot, "Total")
    for cc in range(cLoc, cRowTot + 1):
        cell = ws.cell(rH, cc); cell.font = header; cell.alignment = center; cell.fill = fill_h; cell.border = border

    for i, loc in enumerate(locs):
        r = rFirst + i
        ws.cell(r, cLoc, loc).border = border
        m = meta_by_loc.get(loc, {})
        if include_entity:  ws.cell(r, cEnt, m.get("entity","")).border = border
        if include_address: ws.cell(r, cAdr, m.get("address","")).border = border
        for j, _cov in enumerate(covs):
            cell = ws.cell(r, cFirstCov + j, float(matrix_2d[i][j]) if len(covs) else 0.0)
            cell.style = "Currency2"; cell.border = border
        rc = ws.cell(r, cRowTot, float(row_totals_vec[i]) if i < len(row_totals_vec) else 0.0)
        rc.style = "Currency2"; rc.border = border

    rTotals = rFirst + len(locs)
    ws.cell(rTotals, cLoc, "Total").font = header
    ws.cell(rTotals, cLoc).fill = fill_t; ws.cell(rTotals, cLoc).border = border
    if include_entity:  ws.cell(rTotals, cEnt, "").border = border
    if include_address: ws.cell(rTotals, cAdr, "").border = border

    grand = 0.0
    for j in range(len(covs)):
        v = float(col_totals_vec[j]) if j < len(col_totals_vec) else 0.0
        grand += v
        cell = ws.cell(rTotals, cFirstCov + j, v)
        cell.style = "Currency2"; cell.border = border; cell.font = header; cell.fill = fill_t
    gcell = ws.cell(rTotals, cRowTot, grand); gcell.style = "Currency2"; gcell.border = border; gcell.font = header; gcell.fill = fill_t

    currency_cols = set(range(cFirstCov, cRowTot + 1))
    autosize_columns(ws, range(cLoc, cRowTot + 1), last_row=rTotals, currency_cols=currency_cols)

    rm = wb.create_sheet("READ_ME"); rm["A1"] = "RAS Algorithm Distribution"; rm["A1"].font = header
    rm["A3"] = "IPF (RAS) with exact 2-decimal rounding."

    wb.save(out_path)

# ---------- public entry ----------
def build_ras(path_str: str) -> Path:
    inp = Path(path_str)
    df = load_ras_sheet(inp)
    locs, covs, M, row_vec, col_vec, loc_meta = build_ras_matrix(df)
    out_path = next_output_path(inp.parent)
    write_matrix_generic(out_path, locs, covs, M, row_vec, col_vec, loc_meta)
    return out_path
